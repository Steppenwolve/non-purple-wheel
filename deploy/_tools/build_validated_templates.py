# -*- coding: utf-8 -*-
import pandas as pd, os, re, unicodedata
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import FormulaRule

LAY = r"C:\LOCAL\DESA\non-purple-wheel\non-purple-wheel\layout"
CAT = os.path.join(LAY, "CATALOGOS")
DEPLOY = r"C:\LOCAL\DESA\non-purple-wheel\non-purple-wheel\deploy"
DATA_ROWS = 200
LAST = DATA_ROWS + 1

# ---------- catalog file resolver (name variants) ----------
def norm(s):
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))  # strip accents (ó->o)
    return re.sub(r"[^a-z0-9]", "", s.lower())
_files = [f for f in os.listdir(CAT) if os.path.splitext(f)[1].lower() in (".xlsx", ".xls")]
_fnorm = {}
for f in _files:
    _fnorm.setdefault(norm(os.path.splitext(f)[0]), f)
def resolve(cat):
    k = norm(cat)
    if k in _fnorm: return _fnorm[k]
    cand = [f for kk, f in _fnorm.items() if k in kk or kk in k]
    return cand[0] if cand else None
def safekey(cat):
    return re.sub(r"[^A-Za-z0-9]", "_", cat).upper().strip("_")[:26]
def is_casfim(cat): return norm(cat) in ("ccasfim", "casfim", "casfimv2")

def norm_tipo(t):
    t = str(t).upper().strip()
    if t.startswith("TE"): return "TEXTO"      # handles typo 'TETXO'
    if t.startswith("FEC"): return "FECHA"
    if t.startswith("NUM"): return "NUMERICO"
    if t.startswith("DEC"): return "NUMERICO"  # 'DECIMAL'
    if "AAAA" in t or "AAAAMMDD" in t: return "FECHA"   # date mask used as Tipo (e.g. 'AAAAMMDD')
    return t

def load_catalog(cat, fn):
    df = pd.read_excel(os.path.join(CAT, fn))
    df.columns = [str(c).strip() for c in df.columns]
    clave_col = df.columns[0]
    # If the first column is a NUMERIC code but an alphabetic/ISO column exists, prefer the alpha clave
    # (e.g. MONEDA.xlsx: 'Clave numérica' first, 'Clave Alfabética' is the real 3-char ISO key).
    if "num" in norm(clave_col):
        alt = next((c for c in df.columns[1:] if "alfab" in norm(c) or "iso" in norm(c)), None)
        if alt: clave_col = alt
    desc_col = next((c for c in df.columns if c != clave_col and "descrip" in c.lower()),
                    next((c for c in df.columns if c != clave_col), clave_col))
    out = df[[clave_col, desc_col]].copy(); out.columns = ["Clave", "Descripcion"]
    out["Clave"] = out["Clave"].astype(str).str.strip()
    out = out[~out["Clave"].isin(["", "nan", "NaN", "None"])]
    if is_casfim(cat):
        out["Clave"] = out["Clave"].apply(lambda x: x.zfill(6) if x.isdigit() else x)
    return out.drop_duplicates(subset=["Clave"], keep="first").reset_index(drop=True)

# ---------- styles ----------
HDR_FILL = PatternFill("solid", fgColor="1F4E78"); HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
CAT_HDR_FILL = PatternFill("solid", fgColor="548235")
thin = Side(style="thin", color="BFBFBF"); BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # bgColor set -> paints in CF
RED_FONT = Font(color="9C0006")

def num_format(tipo, long, nombre, cfg):
    if tipo == "FECHA":
        return cfg["date_fechainfo"] if nombre.replace("_", "").upper() == "FECHAINFO" else cfg["date_default"]
    if tipo in ("NUMERICO", "NUMERO"):
        lg = long.replace(" ", "")
        if "," in lg:
            try:
                nd = int(lg.split(",")[1]); return "0." + "0"*nd if nd > 0 else "0"
            except: return "0"
        return "0"
    return "@"

# invalid=TRUE expression (no leading =). nm = named range or None
def invalid_expr(tipo, long, nm, cell):
    if tipo == "FECHA":
        return f'AND({cell}<>"",NOT(ISNUMBER({cell})))'
    if tipo == "TEXTO":
        maxlen = int(long) if str(long).isdigit() else 255
        if nm:
            return f'AND({cell}<>"",OR(LEN({cell})>{maxlen},COUNTIF({nm},{cell})=0))'
        return f'AND({cell}<>"",LEN({cell})>{maxlen})'
    # NUMERICO / NUMERO
    if nm:
        return f'AND({cell}<>"",COUNTIF({nm},{cell})=0)'
    lg = str(long).replace(" ", "")
    if "," in lg:
        n, d = lg.split(","); n = int(n); d = int(d)
        return (f'AND({cell}<>"",IF(ISNUMBER({cell}),'
                f'OR({cell}<>ROUND({cell},{d}),LEN(TEXT(INT(ABS({cell})),"0"))>{n}),TRUE))')
    n = int(lg) if lg.isdigit() else 15
    return (f'AND({cell}<>"",IF(ISNUMBER({cell}),'
            f'OR({cell}<>INT({cell}),LEN(TEXT(ABS({cell}),"0"))>{n}),TRUE))')

def build(cfg):
    lay = pd.read_excel(os.path.join(LAY, cfg["layout"]), sheet_name=cfg["sheet"], header=0)
    lay.columns = [str(c).strip() for c in lay.columns]
    C = cfg["cols"]
    def g(r, key):
        v = r[C[key]]; return "" if pd.isna(v) else str(v).strip()
    fields = []
    for _, r in lay.iterrows():
        if pd.isna(r[C["orden"]]) or not g(r, "nombre"):
            continue
        coment = g(r, "coment") if C.get("coment") else ""   # coment column is optional
        fields.append({"orden": int(r[C["orden"]]), "nombre": g(r, "nombre"), "tipo": norm_tipo(g(r, "tipo")),
                       "long": g(r, "long"), "desc": g(r, "desc"), "coment": coment, "cat": g(r, "cat")})

    # physical column = contiguous 1..N (in ORDEN order); 'orden' stays as layout display value
    fields.sort(key=lambda f: f["orden"])
    for i, f in enumerate(fields, start=1):
        f["col"] = i

    # resolve catalogs (cfg["cat_resolve"] = {catlabel: search-term} overrides auto-resolution)
    override = cfg.get("cat_resolve", {})
    used, missing = [], []
    for f in fields:
        if f["cat"] and f["cat"] not in [u[0] for u in used]:
            fn = resolve(override.get(f["cat"], f["cat"]))
            if fn is None: missing.append(f["cat"])
            else: used.append((f["cat"], fn))
    if missing:
        print(f"  [WARN] {cfg['main']}: catalogos NO encontrados -> {missing}")
    SAFE = {c: safekey(c) for c, _ in used}
    catalogs = {c: load_catalog(c, fn) for c, fn in used}
    # optional per-catalog extra allowed values: cfg["cat_extra"] = {catlabel: [(clave, desc), ...]}
    for cat, extras in cfg.get("cat_extra", {}).items():
        if cat in catalogs:
            add = pd.DataFrame([(str(k), d) for k, d in extras], columns=["Clave", "Descripcion"])
            catalogs[cat] = pd.concat([add, catalogs[cat]], ignore_index=True).drop_duplicates("Clave", keep="first").reset_index(drop=True)
    def nm_of(f):
        return ("CL_" + SAFE[f["cat"]]) if (f["cat"] and f["cat"] in SAFE) else None

    wb = Workbook(); ws = wb.active; ws.title = cfg["main"]

    # catalog sheets (hidden) + named ranges
    for cat, df in catalogs.items():
        sn = ("CAT_" + SAFE[cat])[:31]
        cs = wb.create_sheet(sn)
        cs["A1"] = "Clave"; cs["B1"] = "Descripcion"
        for c in ("A1", "B1"):
            cs[c].font = Font(bold=True, color="FFFFFF", name="Arial", size=10); cs[c].fill = CAT_HDR_FILL
        for i, row in enumerate(df.itertuples(index=False), start=2):
            cs.cell(row=i, column=1, value=str(row.Clave)).number_format = "@"
            cs.cell(row=i, column=2, value=None if pd.isna(row.Descripcion) else str(row.Descripcion))
        cs.column_dimensions["A"].width = 16; cs.column_dimensions["B"].width = 80; cs.sheet_state = "hidden"
        wb.defined_names.add(DefinedName("CL_" + SAFE[cat], attr_text=f"{sn}!$A$2:$A${len(df)+1}"))

    # header + formats + bubbles
    for f in fields:
        L = get_column_letter(f["col"])
        cell = ws.cell(row=1, column=f["col"], value=f["nombre"])
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = BORDER
        lines = [cfg["desc_label"] + ":", f["desc"] or "(sin descripción)"]
        if C.get("coment"):  # only add comment section if the layout has that column
            lines += ["", cfg["com_label"] + ":", f["coment"] or "(sin comentarios)"]
        cmt = Comment("\n".join(lines), "Layout"); cmt.width = 340; cmt.height = 260; cell.comment = cmt
        fmt = num_format(f["tipo"], f["long"], f["nombre"], cfg); width = 14
        if f["tipo"] == "TEXTO" and str(f["long"]).isdigit(): width = max(10, min(30, int(f["long"]) + 2))
        ws.column_dimensions[L].width = width
        for rr in range(2, LAST + 1): ws.cell(row=rr, column=f["col"]).number_format = fmt
    ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 30

    # dropdowns
    for f in fields:
        nm = nm_of(f)
        if not nm: continue
        L = get_column_letter(f["col"])
        dv = DataValidation(type="list", formula1=f"={nm}", allow_blank=True, showDropDown=False)
        dv.showInputMessage = True; dv.showErrorMessage = True
        dv.promptTitle = f["nombre"]; dv.prompt = "Selecciona una Clave del catálogo " + f["cat"]
        dv.errorTitle = "Clave inválida"; dv.error = "El valor debe existir en el catálogo " + f["cat"]
        ws.add_data_validation(dv); dv.add(f"{L}2:{L}{LAST}")

    # conditional formatting (survives paste)
    for f in fields:
        L = get_column_letter(f["col"])
        expr = invalid_expr(f["tipo"], f["long"], nm_of(f), f"{L}2")
        ws.conditional_formatting.add(f"{L}2:{L}{LAST}",
            FormulaRule(formula=[expr], fill=RED_FILL, font=RED_FONT, stopIfTrue=False))

    # hidden CHK sheet
    chk = wb.create_sheet("CHK"); chk.sheet_state = "hidden"
    for f in fields:
        L = get_column_letter(f["col"])
        for r in range(2, LAST + 1):
            expr = invalid_expr(f["tipo"], f["long"], nm_of(f), f"'{cfg['main']}'!{L}{r}")
            chk.cell(row=r, column=f["col"], value=f"=IF({expr},1,0)")

    # VALIDACION report
    vs = wb.create_sheet("VALIDACION")
    vs["A1"] = "RESUMEN DE VALIDACIÓN"; vs["A1"].font = Font(bold=True, size=13)
    vs["A2"] = "Total celdas con error de formato:"; vs["A2"].font = Font(bold=True)
    maxcol = get_column_letter(max(f["col"] for f in fields))
    vs["C2"] = f"=SUM(CHK!A2:{maxcol}{LAST})"
    vs["A3"] = "Estado:"; vs["A3"].font = Font(bold=True)
    vs["C3"] = '=IF(C2=0,"✔ SIN ERRORES DE FORMATO","✖ REVISAR: "&C2&" celda(s) marcada(s) en rojo")'
    vs["C3"].font = Font(bold=True)
    hdr = ["Orden", "Campo", "Tipo", "Long", "Catálogo", "# Errores", "# Vacíos (obligatorio)"]
    hr = 5
    for j, h in enumerate(hdr, start=1):
        c = vs.cell(row=hr, column=j, value=h); c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")
    for i, f in enumerate(fields):
        r = hr + 1 + i; L = get_column_letter(f["col"])
        vs.cell(row=r, column=1, value=f["orden"]); vs.cell(row=r, column=2, value=f["nombre"])
        vs.cell(row=r, column=3, value=f["tipo"]); vs.cell(row=r, column=4, value=f["long"])
        vs.cell(row=r, column=5, value=f["cat"] or "—")
        vs.cell(row=r, column=6, value=f"=SUM(CHK!{L}2:{L}{LAST})")
        vs.cell(row=r, column=7, value=f"=COUNTBLANK('{cfg['main']}'!{L}2:{L}{LAST})")
    last_row = hr + len(fields)
    vs.conditional_formatting.add(f"F{hr+1}:F{last_row}",
        FormulaRule(formula=[f"F{hr+1}>0"], fill=RED_FILL, font=RED_FONT))
    for k, w in {"A": 7, "B": 20, "D": 8, "E": 26, "F": 12, "G": 22}.items():
        vs.column_dimensions[k].width = w
    vs.column_dimensions["C"].width = 42
    wb.move_sheet("VALIDACION", -(len(wb.sheetnames) - 1))

    wb.calculation.fullCalcOnLoad = True
    out_dir = os.path.join(DEPLOY, cfg["out_dir"]); os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, cfg["out_name"])
    wb.save(path)
    print(f"  OK {cfg['main']}: {len(fields)} campos, {len(catalogs)} catalogos -> {path}")

CFGS = [
    {"layout": "Layout_Garantias_V2_FILE_GARANTIAS_IV.xlsx", "sheet": "FILE_GARANTIAS_IV",
     "cols": {"orden": "Orden", "nombre": "Nombre", "tipo": "Tipo", "long": "Longitud",
              "desc": "Descripcion", "coment": "Comentarios", "cat": "Catalogo"},
     "desc_label": "DESCRIPCION", "com_label": "COMENTARIOS",
     "out_dir": "142_ENT_GARANTIAS_IV", "out_name": "PLANTILLA_GARANTIAS_IV.xlsx", "main": "GARANTIAS_IV",
     "date_default": "yyyy-mm-dd", "date_fechainfo": "dd/mm/yyyy"},
    {"layout": "Layout OFF_FX_V11_OFF_CONVAL.xlsx", "sheet": "OFF CONVAL",
     "cols": {"orden": "Orden", "nombre": "Nombre", "tipo": "Tipo", "long": "Formato",
              "desc": "Descripcion", "coment": "Comentario BW", "cat": "Catalogo"},
     "desc_label": "DESCRIPCION", "com_label": "COMENTARIO BW",
     "out_dir": "133_ENT_OFF_CONVAL", "out_name": "PLANTILLA_OFF_CONVAL.xlsx", "main": "OFF_CONVAL",
     "date_default": "yyyy-mm-dd", "date_fechainfo": "dd/mm/yyyy"},
    {"layout": "REPORTOS_MN_ME.xlsx", "sheet": "REPORTOS_MNME",
     "cols": {"orden": "ORDEN", "nombre": "NOMBRE_CAMPO", "tipo": "TIPO DATO", "long": "FORMATO",
              "desc": "DESCRIPCION", "coment": "OBSERVACIONES", "cat": "CATALOGO"},
     "desc_label": "DESCRIPCION", "com_label": "OBSERVACIONES",
     "out_dir": "121_ENT_REPORTOS_MN_ME", "out_name": "PLANTILLA_REPORTOS_MN_ME.xlsx", "main": "REPORTOS_MN_ME",
     "date_default": "yyyy/mm/dd", "date_fechainfo": "yyyy/mm/dd"},
    {"layout": "Layout_Garantias_V2_FILE_GARANTIAS_II.xlsx", "sheet": "FILE_GARANTIAS_II",
     "cols": {"orden": "Orden", "nombre": "Nombre", "tipo": "Tipo", "long": "Longitud",
              "desc": "DESCRIPTION", "coment": "COMMENTS", "cat": "Catalogo"},
     "desc_label": "DESCRIPTION", "com_label": "COMMENTS",
     "out_dir": "140_ENT_GARANTIAS_II", "out_name": "PLANTILLA_GARANTIAS_II.xlsx", "main": "GARANTIAS_II",
     "date_default": "yyyy-mm-dd", "date_fechainfo": "dd/mm/yyyy"},
    {"layout": "Layout_SWAPS_V10_FILE_SWAP_CONVO.xlsx", "sheet": "FILE_SWAP_CONVO",
     "cols": {"orden": "Orden", "nombre": "Nombre", "tipo": "Tipo", "long": "Formato",
              "desc": "Descripcion", "coment": "Validacion", "cat": "Catalogo"},
     "desc_label": "DESCRIPCION", "com_label": "VALIDACION",
     "out_dir": "021_ENT_SWAPS_CONVO", "out_name": "PLANTILLA_SWAP_CONVO.xlsx", "main": "SWAP_CONVO",
     "date_default": "yyyy/mm/dd", "date_fechainfo": "yyyy/mm/dd"},
    {"layout": "Layout_SWAPS_V10_FILE_SWAP.xlsx", "sheet": "FILE_SWAP",
     "cols": {"orden": "Orden", "nombre": "Nombre", "tipo": "Tipo", "long": "formato",
              "desc": "Descripcion", "coment": "Comentario", "cat": "Catalogo"},
     "desc_label": "DESCRIPCION", "com_label": "COMENTARIO",
     "out_dir": "154_ENT_SWAPS", "out_name": "PLANTILLA_SWAP.xlsx", "main": "SWAP",
     "date_default": "yyyy/mm/dd", "date_fechainfo": "yyyy/mm/dd"},
    {"layout": "Layout_REPORTOS_MN_ME_V7.1_2024_CANASTA_REPORTOS.xlsx", "sheet": "CANASTA_REPORTOS",
     "cols": {"orden": "ORDEN", "nombre": "NOMBRE_CAMPO", "tipo": "TIPO DATO", "long": "FORMATO",
              "desc": "DESCRIPCION", "coment": "OBSERVACIONES", "cat": "CATALOGO"},
     "desc_label": "DESCRIPCION", "com_label": "OBSERVACIONES",
     "out_dir": "122_ENT_REPORTOS_CANASTA", "out_name": "PLANTILLA_REPORTOS_CANASTA.xlsx", "main": "REPORTOS_CANASTA",
     "date_default": "yyyy/mm/dd", "date_fechainfo": "yyyy/mm/dd"},
    {"layout": "Layout CF_V3_CF_I.xlsx", "sheet": "CF ESP",
     "cols": {"orden": "ORDEN", "nombre": "NOMBRE DE CAMPO", "tipo": "TIPO DE DATO", "long": "LONGITUD",
              "desc": "DESCRIPCION", "coment": None, "cat": "CATALOGO"},
     "desc_label": "DESCRIPCION", "com_label": "",
     "out_dir": "093_ENT_CF_I", "out_name": "PLANTILLA_CF_I.xlsx", "main": "CF_I",
     "date_default": "dd/mm/yyyy", "date_fechainfo": "dd/mm/yyyy"},
    {"layout": "LAYOUT LID V4 LID CORRESPONSALES.xlsx", "sheet": "LID CORRESPONSALES",
     "cols": {"orden": "ORDEN", "nombre": "NOMBRE_CAMPO", "tipo": "TIPO DATO", "long": "FORMATO",
              "desc": "DESCRIPCION", "coment": "OBSERVACIONES", "cat": "CATALOGO"},
     "desc_label": "DESCRIPCION", "com_label": "OBSERVACIONES",
     "out_dir": "209_LID_CORRESPONSALES", "out_name": "PLANTILLA_LID_CORRESPONSALES.xlsx", "main": "LID_CORRESPONSALES",
     "date_default": "yyyy/mm/dd", "date_fechainfo": "yyyy/mm/dd"},
    {"layout": "LAYOUT LID V4 LID LIQUIDACIONES.xlsx", "sheet": "LID LIQUIDACIONES",
     "cols": {"orden": "ORDEN", "nombre": "NOMBRE_CAMPO", "tipo": "TIPO DATO", "long": "FORMATO",
              "desc": "DESCRIPCION", "coment": None, "cat": "CATALOGO"},
     "desc_label": "DESCRIPCION", "com_label": "",
     "out_dir": "210_LID_LIQUIDACIONES", "out_name": "PLANTILLA_LID_LIQUIDACIONES.xlsx", "main": "LID_LIQUIDACIONES",
     "date_default": "yyyy/mm/dd", "date_fechainfo": "yyyy/mm/dd"},
    {"layout": "LAYOUT LID V4 LID SI.xlsx", "sheet": "LID S1",
     "cols": {"orden": "ORDEN", "nombre": "NOMBRE_CAMPO", "tipo": "TIPO DATO", "long": "FORMATO",
              "desc": "DESCRIPCION", "coment": "OBSERVACIONES", "cat": "CATALOGO"},
     "desc_label": "DESCRIPCION", "com_label": "OBSERVACIONES",
     "out_dir": "211_SECCION_1_LID", "out_name": "PLANTILLA_SECCION_1_LID.xlsx", "main": "SECCION_1_LID",
     "date_default": "yyyy/mm/dd", "date_fechainfo": "yyyy/mm/dd"},
    {"layout": "Layout_SWAPS_V10_FILE_SWAP_FLUJO.xlsx", "sheet": "FILE_SWAP_FLUJO",
     "cols": {"orden": "Orden", "nombre": "Nombre", "tipo": "Tipo", "long": "Formato",
              "desc": "Descripcion", "coment": "Comentarios", "cat": "Catalogo"},
     "desc_label": "DESCRIPCION", "com_label": "COMENTARIOS",
     "out_dir": "157_ENT_SWAP_FLUJOS", "out_name": "PLANTILLA_SWAP_FLUJO.xlsx", "main": "SWAP_FLUJO",
     "date_default": "yyyy/mm/dd", "date_fechainfo": "yyyy/mm/dd"},
    {"layout": "CORE_LayoutTenenciaAdicional_v3.3_Layout_TENENCIA_PRECIO.xlsx", "sheet": "FILE_TENENCIA_PRECIO",
     "cols": {"orden": "Orden", "nombre": "Nombre", "tipo": "Tipo", "long": "Longitud",
              "desc": "Descripcion", "coment": None, "cat": "Catalogo"},
     "desc_label": "DESCRIPCION", "com_label": "",
     "out_dir": "084_ENT_TENENCIA_PRECIO", "out_name": "PLANTILLA_TENENCIA_PRECIO.xlsx", "main": "TENENCIA_PRECIO",
     "date_default": "yyyy/mm/dd", "date_fechainfo": "yyyy/mm/dd"},
    {"layout": "LayoutTenencia_V2_FT_TENENCIA.xlsx", "sheet": "FT_TENENCIA",
     "cols": {"orden": "Orden", "nombre": "Nombre", "tipo": "Tipo", "long": "Formato",
              "desc": "Descripcion", "coment": None, "cat": "Catalogo"},
     "desc_label": "DESCRIPCION", "com_label": "",
     "out_dir": "075_ENT_TENENCIA_REPAD", "out_name": "PLANTILLA_FT_TENENCIA.xlsx", "main": "FT_TENENCIA",
     "date_default": "yyyy/mm/dd", "date_fechainfo": "yyyy/mm/dd",
     "cat_resolve": {"PosicionOperacion": "PosicionOperacion_V4"}},
]

# Para un layout NUEVO: agrega un dict a CFGS con su mapeo de columnas y corre este script.
# Filtro opcional por nombre de 'main' vía argumentos: python build_validated_templates.py SWAP OFF_CONVAL
if __name__ == "__main__":
    import sys
    sel = set(a.upper() for a in sys.argv[1:])
    for cfg in CFGS:
        if sel and cfg["main"].upper() not in sel:
            continue
        try:
            build(cfg)
        except PermissionError:
            print(f"  [LOCKED] {cfg['out_name']} está abierto en Excel — ciérralo y reintento.")
